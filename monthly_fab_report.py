import os
import re
import time
import json
import requests
import feedparser
import urllib.parse
import smtplib
import logging
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse
from dateutil import parser as date_parser
from googlenewsdecoder import gnewsdecoder
from google import genai
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GMAIL_USER = os.getenv("GMAIL_USER")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")

client = genai.Client(api_key=GEMINI_API_KEY)
GEMINI_MODELS = ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.0-flash"]

# =========================================================
# 검색 쿼리 (Fab 전용)
# =========================================================
FAB_QUERIES_EN = [
    "semiconductor fab fabrication plant investment expansion",
    "wafer fab foundry TSMC Intel Micron Samsung construction",
    "chip factory new fab capacity GlobalFoundries SMIC UMC",
    "semiconductor manufacturing facility production line chipmaker",
]
FAB_QUERIES_KO = [
    "반도체 공장 파운드리 팹 투자 생산",
    "삼성전자 SK하이닉스 반도체 공장 생산라인 신규",
]

NEWS_SOURCES = {
    "semiengineering.com": "Semiconductor Engineering",
    "3dincites.com": "3D InCites",
    "semianalysis.com": "SemiAnalysis",
    "digitimes.com": "Digitimes",
    "trendforce.com": "TrendForce",
    "eetimes.com": "EE Times",
    "tomshardware.com": "Tom's Hardware",
    "theregister.com": "The Register",
    "spectrum.ieee.org": "IEEE Spectrum",
    "asia.nikkei.com": "Nikkei Asia",
    "prnewswire.com": "PR Newswire",
    "wccftech.com": "Wccftech",
    "thelec.kr": "TheElec",
    "etnews.com": "ETNews",
    "zdnet.co.kr": "ZDNet Korea",
    "hankyung.com": "Hankyung",
    "sedaily.com": "Seoul Economic Daily",
    "dt.co.kr": "Digital Times",
    "epnc.co.kr": "EPNC",
}


# =========================================================
# 유틸
# =========================================================
def call_gemini(prompt, tag=""):
    for model in GEMINI_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            if resp.text:
                log.info(f"[{tag}] {model} 성공")
                return resp.text
        except Exception as e:
            log.warning(f"[{tag}] {model} 실패: {e}")
            time.sleep(1)
    return None


# =========================================================
# 1. 뉴스 수집
# =========================================================
def fetch_fab_news():
    """지난 30일간 Fab 관련 뉴스 수집."""
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=30)
    raw_articles = []
    seen_links = set()

    def search_rss(query, region, lang):
        encoded = urllib.parse.quote(query)
        url = (
            f"https://news.google.com/rss/search?q={encoded}"
            f"+when:30d&hl={lang}&gl={region}&ceid={region}:{lang}"
        )
        try:
            feed = feedparser.parse(url)
            log.info(f"  RSS ({region}) '{query[:40]}': {len(feed.entries)}건")
            return feed.entries
        except Exception as e:
            log.warning(f"  RSS 실패: {e}")
            return []

    for q in FAB_QUERIES_EN:
        raw_articles.extend(search_rss(q, "US", "en-US"))
        time.sleep(0.5)
    for q in FAB_QUERIES_KO:
        raw_articles.extend(search_rss(q, "KR", "ko"))
        time.sleep(0.5)

    log.info(f"📡 원시 기사 수집: {len(raw_articles)}건 (중복 제거 전)")

    valid = []
    for e in raw_articles:
        link = getattr(e, 'link', None)
        if not link or link in seen_links:
            continue
        seen_links.add(link)

        published = getattr(e, 'published', None)
        if not published:
            continue
        try:
            pub_date = date_parser.parse(published)
            if pub_date.tzinfo is None:
                pub_date = pub_date.replace(tzinfo=timezone.utc)
            if pub_date < cutoff_date:
                continue
        except (ValueError, OverflowError):
            continue

        try:
            decoded = gnewsdecoder(link)
            original_url = (decoded.get('decoded_url', link)
                            if isinstance(decoded, dict) else decoded or link)
        except Exception:
            original_url = link

        original_url = str(original_url)
        domain = urlparse(original_url).netloc.replace("www.", "")
        source_name = next(
            (name for d, name in NEWS_SOURCES.items() if d in domain), domain
        )

        e['display_source'] = source_name
        e['parsed_date'] = pub_date
        e['clean_url'] = original_url
        valid.append(e)

    valid.sort(key=lambda x: x['parsed_date'], reverse=True)
    log.info(f"✅ 유효 기사: {len(valid)}건")
    return valid


# =========================================================
# 2. Gemini 구조화 추출
# =========================================================
def extract_structured_data(articles, idx_offset=0):
    """기사 제목에서 국가·업체·아키텍처 추출 (JSON 반환)."""
    articles_text = ""
    for i, a in enumerate(articles):
        gidx = idx_offset + i + 1
        date_str = a['parsed_date'].strftime("%Y-%m-%d")
        articles_text += f"[{gidx}] {date_str} | {a['display_source']} | {a.get('title', '')}\n"

    prompt = f"""
아래 반도체 뉴스 기사 목록을 분석하여, 각 기사에서 정보를 추출하세요.
반드시 JSON 배열만 출력하세요 (설명 텍스트 없이).

추출 항목:
- idx: 기사 번호 (입력과 동일)
- country: 주요 국가 (한국 / 미국 / 대만 / 중국 / 일본 / 유럽 / 기타 중 하나)
- company: 주요 반도체 업체 (삼성전자 / SK하이닉스 / TSMC / 인텔 / 마이크론 / GlobalFoundries / SMIC / UMC / NVIDIA / AMD / Qualcomm / ASML / 기타 중 하나, 복수면 가장 대표적인 1개)
- architecture: 반도체 아키텍처/공정 (HBM / HBM3E / HBM4 / DRAM / NAND / Logic / FinFET / GAA / 2nm / 3nm / 4nm / 5nm / 7nm / LPDDR / GDDR / CoWoS / 인터포저 / 유리기판 / 기타 중 해당하는 것, 없으면 "해당없음")
- is_fab: Fab/공장/생산시설/공정 직접 관련 여부 (true / false)

[기사 목록]:
{articles_text}

[출력]:
[{{"idx":번호,"country":"...","company":"...","architecture":"...","is_fab":true}},...]
"""

    result = call_gemini(prompt, tag="구조화추출")
    fallback = [
        {"idx": idx_offset + i + 1, "country": "미상", "company": "미상",
         "architecture": "미상", "is_fab": True}
        for i in range(len(articles))
    ]

    if not result:
        return fallback

    match = re.search(r'\[.*\]', result, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            pass

    return fallback


# =========================================================
# 3. 엑셀 생성
# =========================================================
def create_excel(articles, structured_data):
    KST = timezone(timedelta(hours=9))
    now_kst = datetime.now(KST)
    period_start = now_kst - timedelta(days=30)

    filename = f"fab_news_{period_start.strftime('%Y%m')}_{now_kst.strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fab 뉴스 월간 리포트"

    # 스타일 정의
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_fill = PatternFill(start_color="0D3B6E", end_color="0D3B6E", fill_type="solid")
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal="center", vertical="center")

    # 1행: 타이틀
    ws.merge_cells('A1:H1')
    tc = ws['A1']
    tc.value = (
        f"🏭 반도체 Fab 뉴스 월간 리포트  "
        f"({period_start.strftime('%Y년 %m월 %d일')} ~ {now_kst.strftime('%Y년 %m월 %d일')})"
    )
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = title_fill
    tc.alignment = center
    ws.row_dimensions[1].height = 30

    # 2행: 헤더
    headers = ["No.", "날짜", "제목", "출처", "URL", "국가", "주요 업체", "아키텍처/공정"]
    widths = [5, 13, 65, 22, 55, 10, 18, 18]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = center
        c.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 25

    # 3행~: 데이터
    struct_map = {d.get('idx', i + 1): d for i, d in enumerate(structured_data)}

    for i, art in enumerate(articles):
        row = i + 3
        s = struct_map.get(i + 1, {})
        fill = alt_fill if i % 2 == 0 else PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        row_data = [
            i + 1,
            art['parsed_date'].strftime("%Y-%m-%d"),
            art.get('title', ''),
            art.get('display_source', ''),
            art.get('clean_url', ''),
            s.get('country', '미상'),
            s.get('company', '미상'),
            s.get('architecture', '미상'),
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin
            c.fill = fill
            c.font = Font(size=10)

            if col == 1:
                c.alignment = center
            elif col == 3:
                c.alignment = Alignment(vertical="center", wrap_text=True)
            elif col == 5:
                c.hyperlink = val
                c.font = Font(size=10, color="0563C1", underline="single")
                c.alignment = Alignment(vertical="center")
            else:
                c.alignment = center

        ws.row_dimensions[row].height = 38

    # 합계 행
    total_row = len(articles) + 3
    ws.cell(row=total_row, column=1, value="합계").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=2, value=f"총 {len(articles)}건").font = Font(bold=True, size=10)
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = thin
        ws.cell(row=total_row, column=col).alignment = center

    # 열 고정 + 자동 필터
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:H{total_row - 1}"

    wb.save(filename)
    log.info(f"✅ 엑셀 저장: {filename}")
    return filename


# =========================================================
# 4. 이메일 전송 (첨부)
# =========================================================
def send_email_with_attachment(filename, article_count):
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        log.warning("⚠️ 이메일 설정 누락 → 건너뜀")
        return

    KST = timezone(timedelta(hours=9))
    now_kst = datetime.now(KST)
    period_start = now_kst - timedelta(days=30)

    subject = (
        f"🏭 [반도체 Fab 월간 리포트] "
        f"{period_start.strftime('%Y년 %m월')} ({article_count}건)"
    )
    body = (
        f"안녕하세요,\n\n"
        f"{period_start.strftime('%Y년 %m월')} 한 달간의 전세계 반도체 Fab 공장 관련 뉴스를 정리한 월간 리포트를 첨부합니다.\n\n"
        f"📊 수집 기간: {period_start.strftime('%Y-%m-%d')} ~ {now_kst.strftime('%Y-%m-%d')}\n"
        f"📰 총 기사 수: {article_count}건\n"
        f"📋 포함 정보: 날짜 / 출처 / 국가 / 주요 업체 / 반도체 아키텍처·공정\n\n"
        f"첨부 파일을 확인해 주세요.\n\n"
        f"ⓒ 2026 반도체재료개발TFT 김동휘"
    )

    msg = MIMEMultipart()
    msg['From'] = GMAIL_USER
    msg['To'] = "keenhwi@gmail.com"
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    with open(filename, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=15) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            s.send_message(msg)
        log.info("📧 월간 리포트 이메일 전송 성공")
    except Exception as e:
        log.error(f"❌ 이메일 전송 실패: {e}")


# =========================================================
# 5. 메인
# =========================================================
def main():
    log.info("🏭 월간 Fab 뉴스 리포트 시작")

    articles = fetch_fab_news()
    if not articles:
        log.warning("수집된 기사 없음 → 종료")
        return

    log.info("☕ API 보호 대기 (30초)...")
    time.sleep(30)

    # Gemini 구조화 추출 (30건씩 배치)
    BATCH = 30
    all_structured = []
    for start in range(0, len(articles), BATCH):
        batch = articles[start:start + BATCH]
        log.info(f"🔍 구조화 추출 {start + 1}~{start + len(batch)}건...")
        structured = extract_structured_data(batch, idx_offset=start)
        all_structured.extend(structured)
        if start + BATCH < len(articles):
            time.sleep(10)

    filename = create_excel(articles, all_structured)
    send_email_with_attachment(filename, len(articles))
    log.info("✅ 월간 Fab 리포트 완료")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log.critical(f"⚠️ 치명적 에러: {e}", exc_info=True)
        raise SystemExit(1)
